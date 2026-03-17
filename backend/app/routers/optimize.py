from __future__ import annotations

import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models.furnace import OptimizerResult
from app.services.optimizer import run_optimizer

router = APIRouter(prefix="/api", tags=["optimize"])


# ── Request / Response schemas ──────────────────────────────────────────────

class OptimizeRequest(BaseModel):
    upload_id: str = "latest"
    delta_fresh_ethane: float = Field(0.0, description="Extra fresh ethane feed (t/hr)")
    delta_fresh_propane: float = Field(0.0, description="Extra fresh propane feed (t/hr)")
    ethane_feed_purity: float = Field(92.0, description="% ethane in ethane furnace feed")
    propane_feed_purity: float = Field(85.0, description="% propane in propane furnace feed")
    c2_splitter_load: float = Field(82.0, description="Current C2 splitter load %")


# ── POST /api/optimize ──────────────────────────────────────────────────────

@router.post("/optimize")
def optimize_fleet(req: OptimizeRequest, db: Session = Depends(get_db)):
    try:
        result = run_optimizer(
            db,
            upload_id=req.upload_id,
            delta_fresh_ethane=req.delta_fresh_ethane,
            delta_fresh_propane=req.delta_fresh_propane,
            ethane_feed_purity=req.ethane_feed_purity,
            propane_feed_purity=req.propane_feed_purity,
            c2_splitter_load=req.c2_splitter_load,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return result


# ── GET /api/optimize/history ────────────────────────────────────────────────

@router.get("/optimize/history")
def optimizer_history(
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(OptimizerResult)
        .order_by(OptimizerResult.run_at.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "run_id": r.id,
            "snapshot_id": r.snapshot_id,
            "run_at": r.run_at.isoformat() if r.run_at else None,
            "delta_feed_eth": float(r.delta_feed_eth or 0),
            "delta_feed_prop": float(r.delta_feed_prop or 0),
            "ethane_purity": float(r.ethane_purity) if r.ethane_purity else None,
            "propane_purity": float(r.propane_purity) if r.propane_purity else None,
            "gross_profit_M": r.fleet_totals.get("profitGain") if r.fleet_totals else None,
            "net_profit_M": r.fleet_totals.get("netProfit") if r.fleet_totals else None,
        }
        for r in rows
    ]


# ── GET /api/optimize/{run_id}/download ──────────────────────────────────────

@router.get("/optimize/{run_id}/download")
def download_optimizer_result(run_id: int, db: Session = Depends(get_db)):
    record = db.query(OptimizerResult).filter(OptimizerResult.id == run_id).first()
    if not record:
        raise HTTPException(status_code=404, detail=f"Optimizer run {run_id} not found")

    wb = _build_excel(record)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"optimizer_run_{run_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Excel builder ────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 22)


def _build_excel(record: OptimizerResult) -> Workbook:
    wb = Workbook()

    # ── Sheet 1: Per-Furnace Actions ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Per-Furnace Actions"

    headers = [
        "Furnace ID", "Technology", "Feed Type", "Role",
        "Baseline Feed (t/hr)", "Opt Feed (t/hr)", "ΔFeed (t/hr)",
        "ΔCOT (°C)", "ΔSHC",
        "Ethylene Gain (tpy)", "Propylene Gain (tpy)",
        "Profit Gain ($M/yr)", "Uptime Gain (days)",
        "Run Length Δ (days)",
        "Feed Ethane %", "Feed Propane %",
    ]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)
    _style_header(ws1, 1, len(headers))

    per_furnace = record.per_furnace or {}
    row = 2
    for fid in sorted(per_furnace.keys()):
        f = per_furnace[fid]
        vals = [
            fid,
            f.get("technology", ""),
            f.get("feed_type", ""),
            f.get("role", ""),
            f.get("baseline_feed", 0),
            f.get("optFeed", 0),
            f.get("dFeed", 0),
            f.get("dc", 0),
            f.get("ds", 0),
            f.get("ethGain", 0),
            f.get("propGain", 0),
            f.get("profitGain", 0),
            f.get("uptimeGain", 0),
            f.get("runDelta", 0),
            f.get("feed_eth_pct", 0),
            f.get("feed_prop_pct", 0),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=col, value=v)
            cell.border = _THIN_BORDER
            if isinstance(v, float):
                cell.number_format = "#,##0.00"
        row += 1

    _auto_width(ws1)

    # ── Sheet 2: Fleet Totals & Energy ───────────────────────────────────
    ws2 = wb.create_sheet("Fleet Totals")

    totals = record.fleet_totals or {}
    summary_rows = [
        ("Metric", "Value", "Unit"),
        ("Ethylene Gain", totals.get("ethGain", 0), "tpy"),
        ("Propylene Gain", totals.get("propGain", 0), "tpy"),
        ("Gross Profit Gain", totals.get("profitGain", 0), "$M/yr"),
        ("CGC VHP Steam Δ", totals.get("cgc_vhp_delta_tph", 0), "t/hr"),
        ("C2 Splitter VHP Δ", totals.get("c2s_vhp_delta_tph", 0), "t/hr"),
        ("Energy Cost", totals.get("energy_cost_M", 0), "$M/yr"),
        ("Net Profit Gain", totals.get("netProfit", 0), "$M/yr"),
        ("Uptime Gain", totals.get("uptimeGain", 0), "days"),
    ]

    for r_idx, (metric, value, unit) in enumerate(summary_rows, 1):
        ws2.cell(row=r_idx, column=1, value=metric)
        ws2.cell(row=r_idx, column=2, value=value)
        ws2.cell(row=r_idx, column=3, value=unit)
        if r_idx == 1:
            _style_header(ws2, 1, 3)
        else:
            for c in range(1, 4):
                ws2.cell(row=r_idx, column=c).border = _THIN_BORDER
            if isinstance(value, float):
                ws2.cell(row=r_idx, column=2).number_format = "#,##0.000"

    # Optimizer inputs
    config = record.config_used or {}
    ws2.cell(row=len(summary_rows) + 2, column=1, value="Optimizer Inputs").font = Font(bold=True)
    inputs_start = len(summary_rows) + 3
    input_rows = [
        ("Delta Fresh Ethane", float(record.delta_feed_eth or 0), "t/hr"),
        ("Delta Fresh Propane", float(record.delta_feed_prop or 0), "t/hr"),
        ("Ethane Feed Purity", float(record.ethane_purity or 0), "%"),
        ("Propane Feed Purity", float(record.propane_purity or 0), "%"),
        ("C2 Splitter Load", config.get("c2_splitter_load", 0), "%"),
    ]
    for i, (label, val, unit) in enumerate(input_rows):
        ws2.cell(row=inputs_start + i, column=1, value=label)
        ws2.cell(row=inputs_start + i, column=2, value=val)
        ws2.cell(row=inputs_start + i, column=3, value=unit)

    _auto_width(ws2)

    return wb
